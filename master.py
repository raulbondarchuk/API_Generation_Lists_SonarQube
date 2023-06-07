import requests
import base64
import json
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import os
import time # test tiempo de ejecución 
from connectionToken import url, session, headers

def start_timer():
    start_time = time.time()  # start_time
    return start_time

def stop_timer(start_time):
    end_time = time.time()  # end_time
    elapsed_time = end_time - start_time  # Tiempo de ejecución
    return elapsed_time

def get_all_projects():
    page = 1
    projects = []

    while True:
        response = session.get(f"{url}/api/projects/search?", params={"ps": 100, "p": page}, verify=False, timeout=10)
        print(response)
        data = json.loads(response.content.decode('utf-8'))
        projects += data['components']
        if len(data['components']) < 100:
            break
        page += 1
    return projects

projects = get_all_projects() # List of all projects 

# Format Excel file
def auto_size_cells(file_name):
    wb = load_workbook(file_name) 
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = max_length
            sheet.column_dimensions[column].width = adjusted_width

    wb.save(file_name)

# Write DataFrame to Excel file
def writer_save_excel(df, file_name):
    current_dir = os.path.dirname(os.path.abspath(__file__))  # Obtenemos directorio de script
    file_path = os.path.join(current_dir, file_name)  # formamos la ruta completa
    writer = pd.ExcelWriter(file_path)
    df.to_excel(writer, index=False)
    writer._save()
    return writer
    
# Get tags and version from project (True - la respuesta con version, false - la respuesta sin version)
def get_project_tag_and_version(kee, boolVersion):
    if boolVersion:
        project_response = session.get(f"{url}/api/components/show", params={"component": kee}, headers=headers, verify=False, timeout=10)
        project_data = json.loads(project_response.content.decode('utf-8'))
        tags = project_data.get('component', {}).get('tags', [])
        version = project_data.get('component', {}).get('version', '')
        return tags, version
    else:
        project_response = session.get(f"{url}/api/components/show", params={"component": kee}, headers=headers, verify=False, timeout=10)
        project_data = json.loads(project_response.content.decode('utf-8'))
        tags = project_data.get('component', {}).get('tags', [])
        return tags

# Get app_from_tags from project
def get_app_from_tags(tags):
    for tag in tags:
        if tag.startswith('aplic-'):
            app_name = tag[6:].upper().replace('-', '_')
            return f"{app_name}"
    return ''

def get_unanalyzed_projects():
   
    response = session.get(f"{url}/api/projects/search?",params={"onProvisionedOnly": "true","s": "analysisDate","asc": "true"},headers=headers,verify=False,timeout=10)
    projects = json.loads(response.content.decode('utf-8'))['components'] # Una lista de todos los proyectos que tienen ProvisionedOnly
    
    results = []
    for project in projects:
        name = project.get('name', '')
        kee = project.get('key', '')
        tags = get_project_tag_and_version(kee, False) # Get tags from project (sin version)
        #created_at = get_creation_date_from_project(kee) # Get created date from project
        app_from_tags = get_app_from_tags(tags) # Get app_from_tags from project
        results.append({'AppFromTags': app_from_tags, 'Name': name, 'kee': kee, 'tags': ', '.join(tags), }) #'Created_at': created_at

    df = pd.DataFrame(results, columns=['AppFromTags', 'Name', 'kee', 'tags', ]) # 'Created_at'
    df = df.sort_values(['AppFromTags', 'Name', 'kee'], ascending=True, na_position='last')
    
    file_name = 'unanalyzed_projects.xlsx' # File Name
    writer_save_excel(df,file_name) 
    auto_size_cells(file_name)
    #return projects

def get_projects_without_tag_aplic(projects):
    results = []
    for project in projects:
        name = project.get('name', '')
        kee = project.get('key', '')
        tags, version = get_project_tag_and_version(kee, True)
        #created_at = get_creation_date_from_project(kee) 
        last_analisys = project.get('lastAnalysisDate')

        if any(tag.lower().strip().startswith('aplic-') for tag in tags):
            continue 
 
        results.append({'Name': name, 'kee': kee, 'tags': ', '.join(tags),  'last_analisys': last_analisys, 'version': version}) # 'created_at': created_at,

    df = pd.DataFrame(results, columns=['Name', 'kee', 'tags', 'last_analisys', 'version']) #'created_at'
    df = df.sort_values(['Name', 'kee'], ascending=True)

    file_name = 'projects_without_tag_-aplic.xlsx' 
    writer_save_excel(df,file_name) 
    auto_size_cells(file_name)
    #return projects

def get_projects_with_duplicate_name(projects):
    project_names = {}
    for project in projects:
        name = project.get('name', '')  
        if name in project_names:
            project_names[name] += 1
        else:
            project_names[name] = 1

    duplicate_projects = []
    for project in projects:
        name = project.get('name', '')
        kee = project.get('key', '')
        tags, version = get_project_tag_and_version(kee, True)
        #created_at = get_creation_date_from_project(kee) 
        app_from_tags = get_app_from_tags(tags)
        last_analisys = project.get('lastAnalysisDate')

        if project_names[name] > 1:
            duplicate_projects.append({'AppFromTags': app_from_tags, 'Name': name, 'kee': kee, 'tags': ', '.join(tags), 'last_analisys': last_analisys, 'version': version}) # 'created_at': created_at,

    df = pd.DataFrame(duplicate_projects, columns=['AppFromTags', 'Name', 'kee', 'tags', 'last_analisys', 'version']) # 'created_at'
    df = df.sort_values(['AppFromTags', 'Name', 'kee'], ascending=True)

    file_name = 'projects_with_duplicate_names.xlsx'
    writer_save_excel(df,file_name)
    auto_size_cells(file_name)
    #return projects


# Execute
#timer_start = start_timer()
#get_unanalyzed_projects()
#elapsed_time1 = stop_timer(timer_start) # tiempo de ejecución

#timer_start = start_timer()
#get_projects_without_tag_aplic(projects)
#elapsed_time2 = stop_timer(timer_start) # tiempo de ejecución

#timer_start = start_timer()
#get_projects_with_duplicate_name(projects)
#elapsed_time3 = stop_timer(timer_start) # tiempo de ejecución


#print("get_unanalyzed_projects ",elapsed_time1)
#print("get_projects_without_tag_aplic", elapsed_time2)
#print("get_projects_with_duplicate_name", elapsed_time3)